import streamlit as st
import pandas as pd
from io import BytesIO
import warnings
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter

# Suppress the specific UserWarning from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Data cleaning function with correct calculations for CTR and ACOS
def clean_amazon_data(file):
    amazon_data = pd.read_excel(file)
    amazon_data.columns = amazon_data.columns.str.strip()
    amazon_data['Date'] = pd.to_datetime(amazon_data['Date'])
    amazon_data.fillna(0, inplace=True)

    columns_to_remove = [
        'Top-of-search Impression Share', 
        'Total Return on Advertising Spend (ROAS)', '7 Day Total Orders (#)',
        '7 Day Conversion Rate', '7 Day Advertised SKU Units (#)',
        '7 Day Other SKU Units (#)', '7 Day Advertised SKU Sales', '7 Day Other SKU Sales',
        'Currency', 'Ad Group Name','Total Advertising Cost of Sales (ACOS)', 'Click-Thru Rate (CTR)'
    ]
    amazon_data.drop(columns=columns_to_remove, inplace=True)

    amazon_data.rename(columns={
        '7 Day Total Sales': 'Ad Sales', 
        'Spend': 'Ad Spend', 
        '7 Day Total Units (#)': 'Units',
        'Cost Per Click (CPC)': 'CPC',
        'Portfolio name': 'Portfolio'
    }, inplace=True)

    numeric_columns = ['Impressions', 'Ad Sales', 'Units', 'Clicks', 'Ad Spend', 'CPC']
    amazon_data[numeric_columns] = amazon_data[numeric_columns].apply(pd.to_numeric, errors='coerce').round(2)

    # Calculate CTR and ACOS as numeric values
    amazon_data['CTR'] = (amazon_data['Clicks'] / amazon_data['Impressions']).round(4)
    amazon_data['ACOS'] = (amazon_data['Ad Spend'] / amazon_data['Ad Sales']).round(2)

    amazon_data = amazon_data[amazon_data['Ad Spend'] > 0]

    exclude_targeting = ['loose-match', 'close-match', 'complements', 'substitutes']
    amazon_data = amazon_data[~amazon_data['Targeting'].isin(exclude_targeting)]

    exclude_keywords = ['category', 'B0']
    pattern = '|'.join(exclude_keywords)
    amazon_data = amazon_data[~amazon_data['Targeting'].str.contains(pattern, case=False, na=False)]

    amazon_data.reset_index(drop=True, inplace=True)

    return amazon_data

# Function to create the 'Review' sheet
def create_review_sheet(cleaned_data):
    end_date = cleaned_data['Date'].max()
    start_date = end_date - timedelta(weeks=6)
    recent_data = cleaned_data[(cleaned_data['Date'] > start_date) & (cleaned_data['Date'] <= end_date)]
    
    ad_spend = recent_data.groupby(['Portfolio', pd.Grouper(key='Date', freq='W-SUN')])['Ad Spend'].sum().unstack(fill_value=0)
    ad_sales = recent_data.groupby(['Portfolio', pd.Grouper(key='Date', freq='W-SUN')])['Ad Sales'].sum().unstack(fill_value=0)
    acos = (ad_spend / ad_sales).replace([float('inf'), -float('inf'), pd.NA], 0).round(2)
    
    review_data = pd.concat([ad_spend, acos], axis=1, keys=['Spend', 'ACOS'])
    review_data.columns = [f"{col[1].strftime('%m-%d-%Y')} {col[0]}" for col in review_data.columns]
    
    columns = ['Portfolio']
    dates = sorted(set(col.split()[0] for col in review_data.columns))
    for date in dates:
        columns.append(f"{date} Spend")
        columns.append(f"{date} ACOS")
    review_data = review_data.reset_index()
    review_data = review_data[columns]
    
    last_week_col = [col for col in review_data.columns if 'Spend' in col][-1]
    review_data.sort_values(by=last_week_col, ascending=False, inplace=True)
    
    return review_data

# Function to create individual portfolio sheets with Ad Sales, Ad Spend, and ACOS columns
def create_portfolio_sheets(cleaned_data):
    portfolio_sheets = {}

    end_date = cleaned_data['Date'].max()
    start_date = end_date - timedelta(weeks=6)
    recent_data = cleaned_data[(cleaned_data['Date'] > start_date) & (cleaned_data['Date'] <= end_date)]

    portfolios = recent_data['Portfolio'].unique()

    for portfolio in portfolios:
        portfolio_data = recent_data[recent_data['Portfolio'] == portfolio]

        portfolio_sheet = portfolio_data.groupby(['Targeting', pd.Grouper(key='Date', freq='W-SUN')]).agg(
            Ad_Sales=('Ad Sales', 'sum'),
            Ad_Spend=('Ad Spend', 'sum'),
            ACOS=('ACOS', 'mean')  # Assuming you want to average the ACOS over the week
        ).unstack(fill_value=0)

        portfolio_sheet.columns = ['_'.join([col[0], col[1].strftime('%m-%d-%Y')]) for col in portfolio_sheet.columns]

        portfolio_sheet.reset_index(inplace=True)

        # Adding debug print statements
        print(f"Columns before processing for portfolio {portfolio}: {portfolio_sheet.columns}")

        for date in sorted(set(col.split('_')[1] for col in portfolio_sheet.columns if '_' in col)):
            spend_col = f'Ad_Spend_{date}'
            sales_col = f'Ad_Sales_{date}'
            acos_col = f'ACOS_{date}'

            # Check if columns exist before using them
            if spend_col in portfolio_sheet.columns and sales_col in portfolio_sheet.columns:
                portfolio_sheet[f'Spend_{date}'] = portfolio_sheet[spend_col]
                portfolio_sheet[f'ACOS_{date}'] = portfolio_sheet[acos_col]
            else:
                print(f"Missing column for date {date}: {spend_col} or {sales_col}")

        columns_to_keep = ['Targeting'] + [col for col in portfolio_sheet.columns if 'Ad_Sales' in col or 'Spend' in col or 'ACOS' in col]
        portfolio_sheet = portfolio_sheet[columns_to_keep]

        # Print columns after processing to debug
        print(f"Columns after processing for portfolio {portfolio}: {portfolio_sheet.columns}")

        portfolio_sheets[portfolio] = portfolio_sheet

    return portfolio_sheets

# Function to convert DataFrame to Excel in memory and adjust column widths
def to_excel(cleaned_data, review_data, portfolio_sheets):
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    review_data.to_excel(writer, index=False, sheet_name='Review')
    cleaned_data.to_excel(writer, index=False, sheet_name='Base')

    for portfolio, data in portfolio_sheets.items():
        data.to_excel(writer, index=False, sheet_name=portfolio)

    workbook = writer.book

    currency_style = NamedStyle(name="currency_style", number_format="$#,##0.00")
    percentage_style = NamedStyle(name="percentage_style", number_format="0.00%")
    number_style = NamedStyle(name="number_style", number_format="#,##0")
    decimal_style = NamedStyle(name="decimal_style", number_format="0.00")

    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.coordinate in worksheet.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        if sheetname == 'Review':
            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if 'Spend' in cell.coordinate:
                        cell.number_format = currency_style.number_format
                    elif 'ACOS' in cell.coordinate:
                        cell.number_format = decimal_style.number_format
        elif sheetname == 'Base':
            for col_num, column in enumerate(worksheet.iter_cols(min_row=2, min_col=1), 1):
                col_letter = get_column_letter(col_num)
                header_value = worksheet.cell(row=1, column=col_num).value
                if header_value in ['CTR', 'ACOS']:
                    for cell in column:
                        cell.number_format = percentage_style.number_format  # Apply decimal format with two decimal places
                elif header_value in ['Impressions', 'Clicks', 'Units']:
                    for cell in column:
                        cell.number_format = number_style.number_format
                elif header_value in ['Ad Spend', 'Ad Sales']:
                    for cell in column:
                        cell.number_format = currency_style.number_format

    writer.close()
    processed_data = output.getvalue()
    return processed_data

# Streamlit app
def main():
    st.title("AMAZON DATA: TARGETING REVIEW")

    uploaded_file = st.file_uploader("Upload Amazon SP Targeting Report (Make sure the time unit is Daily)", type="xlsx")

    if uploaded_file is not None:
        st.write("File uploaded successfully!")

        cleaned_data = clean_amazon_data(uploaded_file)
        review_data = create_review_sheet(cleaned_data)
        portfolio_sheets = create_portfolio_sheets(cleaned_data)

        #st.write("Preview of cleaned data:")
        #st.dataframe(cleaned_data.head())

        #st.write("Preview of review data:")
        #st.dataframe(review_data.head())

        for portfolio, data in portfolio_sheets.items():
            st.write(f"\nPreview of portfolio data for {portfolio}:")
            st.dataframe(data.head())

        cleaned_data_excel = to_excel(cleaned_data, review_data, portfolio_sheets)

        st.download_button(
            label="Download XLSX file",
            data=cleaned_data_excel,
            file_name='Target_Review.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

if __name__ == "__main__":
    main()
