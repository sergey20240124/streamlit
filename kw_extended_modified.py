import streamlit as st
import pandas as pd
from io import BytesIO
import warnings
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment
from openpyxl.utils import get_column_letter

# Suppress the specific UserWarning from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Data cleaning function with correct calculations for CTR and ACOS
def clean_amazon_data(file):
    try:
        amazon_data = pd.read_excel(file)
        amazon_data.columns = amazon_data.columns.str.strip()
        amazon_data['Date'] = pd.to_datetime(amazon_data['Date'])
        amazon_data.fillna(0, inplace=True)

        columns_to_remove = [
            'Top-of-search Impression Share', 
            'Total Return on Advertising Spend (ROAS)', '7 Day Total Orders (#)',
            '7 Day Conversion Rate', '7 Day Advertised SKU Units (#)',
            '7 Day Other SKU Units (#)', '7 Day Advertised SKU Sales', '7 Day Other SKU Sales',
            'Currency', 'Ad Group Name', 'Total Advertising Cost of Sales (ACOS)', 'Click-Thru Rate (CTR)'
        ]
        amazon_data.drop(columns=columns_to_remove, inplace=True)

        amazon_data.rename(columns={
            '7 Day Total Sales': 'Ad Sales', 
            'Spend': 'Ad Spend', 
            '7 Day Total Units (#)': 'Units',
            'Cost Per Click (CPC)': 'CPC',
            'Portfolio name': 'Portfolio'
        }, inplace=True)

        numeric_columns = ['Impressions', 'Ad Sales', 'Units', 'Clicks', 'Ad Spend', 'CPC']
        amazon_data[numeric_columns] = amazon_data[numeric_columns].apply(pd.to_numeric, errors='coerce').round(2)

        # Calculate CTR and ACOS as numeric values
        amazon_data['CTR'] = (amazon_data['Clicks'] / amazon_data['Impressions']).round(4)
        amazon_data['ACOS'] = (amazon_data['Ad Spend'] / amazon_data['Ad Sales']).round(2)

        amazon_data = amazon_data[amazon_data['Ad Spend'] > 0]

        exclude_targeting = ['loose-match', 'close-match', 'complements', 'substitutes']
        amazon_data = amazon_data[~amazon_data['Targeting'].isin(exclude_targeting)]

        exclude_keywords = ['category', 'B0']
        pattern = '|'.join(exclude_keywords)
        amazon_data = amazon_data[~amazon_data['Targeting'].str.contains(pattern, case=False, na=False)]

        amazon_data.reset_index(drop=True, inplace=True)

        print("Cleaned data:")
        print(amazon_data.head())
        
        return amazon_data

    except Exception as e:
        st.error(f"Error in cleaning data: {e}")
        print(f"Error in cleaning data: {e}")
        return pd.DataFrame()

# Function to create the 'Review' sheet
def create_review_sheet(cleaned_data):
    try:
        end_date = cleaned_data['Date'].max()
        start_date = end_date - timedelta(weeks=6)
        recent_data = cleaned_data[(cleaned_data['Date'] > start_date) & (cleaned_data['Date'] <= end_date)]
        
        ad_spend = recent_data.groupby(['Portfolio', pd.Grouper(key='Date', freq='W-SUN')])['Ad Spend'].sum().unstack(fill_value=0)
        ad_sales = recent_data.groupby(['Portfolio', pd.Grouper(key='Date', freq='W-SUN')])['Ad Sales'].sum().unstack(fill_value=0)
        acos = (ad_spend / ad_sales).replace([float('inf'), -float('inf'), pd.NA], 0).round(2)
        
        review_data = pd.concat([ad_spend, acos], axis=1, keys=['Spend', 'ACOS'])
        review_data.columns = [f"{col[1].strftime('%m-%d-%Y')} {col[0]}" for col in review_data.columns]
        
        columns = ['Portfolio']
        dates = sorted(set(col.split()[0] for col in review_data.columns))
        for date in dates:
            columns.append(f"{date} Spend")
            columns.append(f"{date} ACOS")
        review_data = review_data.reset_index()
        review_data = review_data[columns]
        
        last_week_col = [col for col in review_data.columns if 'Spend' in col][-1]
        review_data.sort_values(by=last_week_col, ascending=False, inplace=True)

        print("Review data:")
        print(review_data.head())
        
        return review_data

    except Exception as e:
        st.error(f"Error in creating review sheet: {e}")
        print(f"Error in creating review sheet: {e}")
        return pd.DataFrame()

# Updated function to create individual portfolio sheets with Ad Sales, Ad Spend, and ACOS columns
def create_portfolio_sheets(cleaned_data):
    try:
        portfolio_sheets = {}

        end_date = cleaned_data['Date'].max()
        start_date = end_date - timedelta(weeks=6)
        recent_data = cleaned_data[(cleaned_data['Date'] > start_date) & (cleaned_data['Date'] <= end_date)]

        portfolios = recent_data['Portfolio'].unique()

        for portfolio in portfolios:
            try:
                portfolio_data = recent_data[recent_data['Portfolio'] == portfolio]

                portfolio_sheet = portfolio_data.groupby(['Targeting', pd.Grouper(key='Date', freq='W-SUN')]).agg(
                    Ad_Sales=('Ad Sales', 'sum'),
                    Ad_Spend=('Ad Spend', 'sum'),
                    ACOS=('ACOS', 'mean')  # Assuming you want to average the ACOS over the week
                ).unstack(fill_value=0)

                portfolio_sheet.columns = ['_'.join([col[0], col[1].strftime('%m-%d-%Y')]) for col in portfolio_sheet.columns]
                portfolio_sheet.reset_index(inplace=True)

                for date in sorted(set(col.split('_')[1] for col in portfolio_sheet.columns if '_' in col)):
                    spend_col = f'Ad_Spend_{date}'
                    sales_col = f'Ad_Sales_{date}'
                    acos_col = f'ACOS_{date}'

                    if spend_col in portfolio_sheet.columns and sales_col in portfolio_sheet.columns:
                        portfolio_sheet[f'Spend_{date}'] = portfolio_sheet[spend_col]
                        portfolio_sheet[f'ACOS_{date}'] = portfolio_sheet[acos_col]
                    else:
                        print(f"Missing column for date {date}: {spend_col} or {sales_col}")

                columns_to_keep = ['Targeting'] + [col for col in portfolio_sheet.columns if 'Ad_Sales' in col or 'Spend' in col or 'ACOS' in col]
                portfolio_sheet = portfolio_sheet[columns_to_keep]

                print(f"Columns after processing for portfolio {portfolio}: {portfolio_sheet.columns}")

                portfolio_sheets[portfolio] = portfolio_sheet
            except Exception as e:
                print(f"Error processing portfolio {portfolio}: {e}")

        return portfolio_sheets

    except Exception as e:
        st.error(f"Error in creating portfolio sheets: {e}")
        print(f"Error in creating portfolio sheets: {e}")
        return {}

# Updated function to create the 'Spend Tracking' sheet with adjusted calculations
def create_spend_tracking_sheet(cleaned_data):
    try:
        end_date = cleaned_data['Date'].max()
        start_date = end_date - timedelta(weeks=5)
        last_week_start = end_date - timedelta(weeks=1)
        last_week_end = end_date

        # Data for the prior 4 weeks excluding the last week
        prior_4_weeks_data = cleaned_data[(cleaned_data['Date'] > start_date) & (cleaned_data['Date'] <= last_week_start)]
        # Data for the last week
        last_week_data = cleaned_data[(cleaned_data['Date'] > last_week_start) & (cleaned_data['Date'] <= last_week_end)]

        # Add a new column to indicate the week
        prior_4_weeks_data['Week'] = prior_4_weeks_data['Date'].dt.isocalendar().week
        last_week_data['Week'] = last_week_data['Date'].dt.isocalendar().week

        # Group by Portfolio, Campaign Name, Targeting, and Week to calculate total spend and count of weeks
        spend_by_keyword = prior_4_weeks_data.groupby(['Portfolio', 'Campaign Name', 'Targeting', 'Week']).agg(
            weekly_spend=('Ad Spend', 'sum')
        ).reset_index()

        # Calculate the number of weeks active and the total spend per keyword
        spend_summary = spend_by_keyword.groupby(['Portfolio', 'Campaign Name', 'Targeting']).agg(
            total_spend=('weekly_spend', 'sum'),
            weeks_active=('Week', 'count')
        ).reset_index()

        # Calculate the average spend by dividing the total spend by the number of weeks active
        spend_summary['4 Week Avg Spend'] = spend_summary['total_spend'] / spend_summary['weeks_active']

        # Calculate the last week's spend
        last_week_spend = last_week_data.groupby(['Portfolio', 'Campaign Name', 'Targeting'])['Ad Spend'].sum().reset_index()

        # Merge the average spend and last week's spend into a single DataFrame
        spend_tracking = spend_summary.merge(
            last_week_spend.rename(columns={'Ad Spend': 'Last Week Spend'}),
            on=['Portfolio', 'Campaign Name', 'Targeting'],
            how='left'
        ).fillna(0)

        # Calculate the change percentage
        spend_tracking['Change'] = ((spend_tracking['Last Week Spend'] - spend_tracking['4 Week Avg Spend']) / spend_tracking['4 Week Avg Spend']).round(2)

        # Determine the status based on the change percentage
        spend_tracking['Status'] = spend_tracking['Change'].apply(
            lambda x: 'Increase 25% plus' if x > 0.25 else ('Decrease 25% plus' if x < -0.25 else 'Stable')
        )

        # Filter out 'Stable' keywords and those with zero spend last week
        spend_tracking = spend_tracking[(spend_tracking['Status'] != 'Stable') & (spend_tracking['Last Week Spend'] > 0)]

        # Select and order the final columns
        spend_tracking = spend_tracking[['Portfolio', 'Campaign Name', 'Targeting', '4 Week Avg Spend', 'Last Week Spend', 'Status']]

        print("Spend tracking data:")
        print(spend_tracking.head())

        return spend_tracking

    except Exception as e:
        st.error(f"Error in creating spend tracking sheet: {e}")
        print(f"Error in creating spend tracking sheet: {e}")
        return pd.DataFrame()




# Function to convert DataFrame to Excel in memory and adjust column widths
def to_excel(cleaned_data, review_data, portfolio_sheets, spend_tracking_data):
    try:
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')

        spend_tracking_data.to_excel(writer, index=False, sheet_name='Spend Tracking')
        review_data.to_excel(writer, index=False, sheet_name='Review')
        cleaned_data.to_excel(writer, index=False, sheet_name='Base')

        for portfolio, data in portfolio_sheets.items():
            data.to_excel(writer, index=False, sheet_name=portfolio)

        workbook = writer.book

        currency_style = NamedStyle(name="currency_style", number_format="$#,##0.00")
        percentage_style = NamedStyle(name="percentage_style", number_format="0.00%")
        number_style = NamedStyle(name="number_style", number_format="#,##0")
        decimal_style = NamedStyle(name="decimal_style", number_format="0.00")

        for sheetname in workbook.sheetnames:
            worksheet = workbook[sheetname]

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.coordinate in worksheet.merged_cells:
                        continue
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            if sheetname == 'Spend Tracking':
                for col_num, column in enumerate(worksheet.iter_cols(min_row=2, min_col=1), 1):
                    col_letter = get_column_letter(col_num)
                    header_value = worksheet.cell(row=1, column=col_num).value
                    if header_value in ['4 Week Avg Spend', 'Last Week Spend']:
                        for cell in column:
                            cell.number_format = currency_style.number_format
                    elif header_value == 'Status':
                        for cell in column:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

        writer.close()
        processed_data = output.getvalue()

        print("Excel data prepared for download.")
        
        return processed_data

    except Exception as e:
        st.error(f"Error in converting to Excel: {e}")
        print(f"Error in converting to Excel: {e}")
        return None

# Streamlit app
def main():
    st.title("AMAZON DATA: TARGETING REVIEW")

    uploaded_file = st.file_uploader("Upload Amazon SP Targeting Report (Make sure the time unit is Daily)", type="xlsx")

    if uploaded_file is not None:
        st.write("File uploaded successfully!")

        cleaned_data = clean_amazon_data(uploaded_file)
        review_data = create_review_sheet(cleaned_data)
        portfolio_sheets = create_portfolio_sheets(cleaned_data)
        spend_tracking_data = create_spend_tracking_sheet(cleaned_data)

        st.write("Preview of review data:")
        st.dataframe(review_data.head())

        st.write("Preview of spend tracking data:")
        st.dataframe(spend_tracking_data.head())

        cleaned_data_excel = to_excel(cleaned_data, review_data, portfolio_sheets, spend_tracking_data)

        if cleaned_data_excel:
            st.download_button(
                label="Download XLSX file",
                data=cleaned_data_excel,
                file_name='Target_Review.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

if __name__ == "__main__":
    main()
